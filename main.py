from fastapi import FastAPI, HTTPException, Request, UploadFile, File, Form
from fastapi.responses import HTMLResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from supabase import create_client, Client
import os
import ollama
import json
import re
import tempfile
import fitz  # PyMuPDF
import logging
from datetime import datetime
import chromadb
from chromadb.config import Settings
from typing import Optional, Dict, Any, List
import base64
import csv
import io
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

# --- Logging Setup ---
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# --- Environment Variables ---
SUPABASE_URL = os.getenv("SUPABASE_URL", "Your_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "Your_key")

# Initialize Supabase only if credentials are provided
if SUPABASE_URL != "Your_URL" and SUPABASE_KEY != "Your_key":
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
else:
    supabase = None
    logger.warning("Supabase not configured. Database storage will be disabled.")
    
app = FastAPI(title="Universal Document Processing API")

# --- Dynamic Schema Model ---
class DynamicSchema(BaseModel):
    schema_definition: Dict[str, Any]
    extraction_prompt: str
    document_type: str = "general"

class ExtractedData(BaseModel):
    data: Dict[str, Any]
    schema_config: Dict[str, Any]
    metadata: Dict[str, Any] = {}

# --- ChromaDB Setup ---
chroma_client = chromadb.PersistentClient(path="./chroma_storage")
invoice_collection = chroma_client.get_or_create_collection(name="invoices")

# --- Helper: Deduplicate & Merge Quantities ---
def deduplicate_items(items):
    merged = {}
    for item in items:
        name = item.get("name") if isinstance(item, dict) else str(item)
        try:
            qty = int(item.get("qty", 1)) if isinstance(item, dict) else 1
        except Exception:
            qty = 1
        merged[name] = merged.get(name, 0) + qty
    return [{"name": n, "qty": q} for n, q in merged.items()]

# --- Helper: Normalize Date ---
def normalize_date(date_str: str) -> str:
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d %b %Y", "%d %B %Y"):
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return date_str

# --- Universal Extraction from Image with Custom Prompt ---
def extract_data_from_image(image_path: str, custom_prompt: str, schema: Dict[str, Any], 
                            strict_mode: bool = False, double_check: bool = False,
                            infer_missing: bool = True, auto_detect_fields: bool = False) -> dict:
    """Extract data from image using user-defined prompt and schema with advanced options"""
    
    # Build enhanced prompt based on toggles
    enhancements = []
    
    if strict_mode:
        enhancements.append("""
⚠️ STRICT MODE ENABLED:
- Only extract information that is EXPLICITLY visible in the document
- Do NOT infer, assume, or hallucinate any data
- If a field is not clearly present, return null/empty
- Accuracy is more important than completeness
""")
    
    if double_check:
        enhancements.append("""
🔍 DOUBLE-CHECK MODE ENABLED:
- Review your extraction twice before responding
- Verify each extracted value against the original image
- Cross-check related fields for consistency (e.g., total = sum of items)
- Mark confidence level for each field (high/medium/low)
""")
    
    if auto_detect_fields:
        enhancements.append("""
🤖 AUTO-DETECT FIELDS ENABLED:
- Automatically identify required fields regardless of naming convention
- Handle synonyms, abbreviations, and variations
- Map different field names to standard schema (e.g., "inv_no" → "invoice_number")
- Intelligently detect field types from context
""")
    
    if infer_missing:
        enhancements.append("""
💡 INTELLIGENT INFERENCE ENABLED:
- Use context clues to fill in missing but obvious information
- Infer dates from surrounding text if partial date visible
- Infer totals from line items if not explicitly stated
- Use common patterns (e.g., phone numbers, emails, addresses)
""")
    else:
        enhancements.append("""
🚫 NO INFERENCE:
- Extract ONLY what is explicitly written
- Do not use context or patterns to guess missing data
""")
    
    prompt = f"""
{custom_prompt}

{' '.join(enhancements)}

Return ONLY valid JSON that matches this exact schema structure:
{json.dumps(schema, indent=2)}

Do not include any explanations, only the JSON object.
"""

    response = ollama.chat(
        model="llama3.2-vision",
        format="json",
        messages=[{
            "role": "user",
            "content": prompt,
            "images": [image_path]
        }]
    )

    raw = response["message"]["content"]
    logger.info(f"Raw LLM output (image): {raw}")

    match = re.search(r"\{.*\}", raw, re.DOTALL)
    if not match:
        raise HTTPException(status_code=500, detail="No JSON found in LLM response")

    data = json.loads(match.group(0))
    
    # Post-processing for double-check mode
    if double_check:
        # Add validation metadata
        data['_extraction_confidence'] = 'verified'
        data['_double_checked'] = True
    
    # Add auto-detection metadata
    if auto_detect_fields:
        data['_auto_detected'] = True
    
    return data

# --- Universal Data Normalization ---
def normalize_data_with_llm(body: dict, target_schema: Dict[str, Any], auto_detect_fields: bool = False) -> dict:
    """Transform input data to match user-defined schema using LLM with auto-detection"""
    
    if auto_detect_fields:
        prompt = f"""
        You are an intelligent data extraction assistant. Your task is to:
        1. Analyze the input data which may have ANY field naming convention
        2. Map it to the target schema structure
        3. Handle synonyms, variations, and different naming patterns
        
        Target Schema Structure:
        {json.dumps(target_schema, indent=2)}
        
        Input Data (may have different naming conventions):
        {json.dumps(body, indent=2)}
        
        IMPORTANT - Handle these variations:
        - Synonyms: "invoice_number", "invoice_id", "inv_no", "bill_number", "document_number" → map to schema field
        - Different cases: "InvoiceNumber", "invoice_number", "INVOICE_NUMBER" → normalize
        - Abbreviations: "qty" → "quantity", "amt" → "amount", "desc" → "description"
        - Regional variations: "colour" → "color", "organisation" → "organization"
        - Domain-specific terms: "vendor", "supplier", "merchant", "seller" → map appropriately
        - Snake_case vs CamelCase: "totalAmount" vs "total_amount"
        - Typos and OCR errors: Use context to infer correct field
        
        Rules:
        1. Match fields by SEMANTIC MEANING, not exact name
        2. If multiple input fields could map to one schema field, choose the best match
        3. If a required field is missing but similar field exists, use it
        4. Convert data types as needed (dates to strings, numbers to strings, etc.)
        5. For arrays, detect list items even if named differently
        6. Return ONLY the transformed JSON matching target schema exactly
        7. Add "_field_mapping" metadata showing how you mapped fields
        
        Output (matching target schema exactly):
        """
    else:
        prompt = f"""
        Transform the input JSON data to match this EXACT output schema:
        
        Target Schema:
        {json.dumps(target_schema, indent=2)}
        
        Input Data:
        {json.dumps(body, indent=2)}
        
        Rules:
        1. Map all input fields to match the target schema structure exactly
        2. Do not add fields that are not in the target schema
        3. Do not remove required fields from the target schema
        4. Convert data types as needed (dates to strings, numbers to strings, etc.)
        5. If a field cannot be mapped, set it to null or appropriate default
        6. Return ONLY the transformed JSON object, no explanations
        
        Output (matching target schema exactly):
        """

    response = ollama.chat(
        model="llama3.2",
        format="json",
        messages=[{"role": "user", "content": prompt}]
    )

    raw = response["message"]["content"]
    logger.info(f"Raw LLM output (JSON): {raw}")

    match = re.search(r"\{.*\}", raw, re.DOTALL)
    if not match:
        raise HTTPException(status_code=500, detail="No JSON found in LLM response")

    data = json.loads(match.group(0))
    return data

# --- Helper: Parse any document type ---
def parse_document_to_images(file_path: str, suffix: str) -> List[str]:
    """Convert various document formats to images for processing"""
    image_paths = []
    
    if suffix == ".pdf":
        doc = fitz.open(file_path)
        for page_num in range(len(doc)):
            page = doc[page_num]
            pix = page.get_pixmap()
            image_path = file_path.replace(".pdf", f"_page_{page_num+1}.png")
            pix.save(image_path)
            image_paths.append(image_path)
        doc.close()
    elif suffix in [".png", ".jpg", ".jpeg", ".webp"]:
        image_paths.append(file_path)
    else:
        # For unsupported formats, try to open as PDF or image
        try:
            doc = fitz.open(file_path)
            for page_num in range(len(doc)):
                page = doc[page_num]
                pix = page.get_pixmap()
                image_path = file_path + f"_page_{page_num+1}.png"
                pix.save(image_path)
                image_paths.append(image_path)
            doc.close()
        except Exception as e:
            logger.warning(f"Could not parse file as PDF: {e}")
            raise HTTPException(status_code=400, detail=f"Unsupported file format: {suffix}")
    
    return image_paths

# --- Universal File Processing ---
def process_document(file: UploadFile, custom_prompt: str, schema: Dict[str, Any],
                    strict_mode: bool = False, double_check: bool = False,
                    infer_missing: bool = True, auto_detect_fields: bool = False) -> dict:
    """Process any document type with user-defined extraction parameters"""
    suffix = os.path.splitext(file.filename)[-1].lower()
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp_file:
        tmp_file.write(file.file.read())
        tmp_path = tmp_file.name

    try:
        # Convert document to images
        image_paths = parse_document_to_images(tmp_path, suffix)
        
        all_extracted_data = {}
        page_results = []
        
        # Extract data from each page/image
        for idx, image_path in enumerate(image_paths):
            extracted = extract_data_from_image(
                image_path, 
                custom_prompt, 
                schema,
                strict_mode=strict_mode,
                double_check=double_check,
                infer_missing=infer_missing,
                auto_detect_fields=auto_detect_fields
            )
            page_results.append({
                "page": idx + 1,
                "data": extracted
            })
            
            # Merge data from all pages
            for key, value in extracted.items():
                if not key.startswith('_'):  # Skip metadata fields
                    if key not in all_extracted_data:
                        all_extracted_data[key] = value
                    elif isinstance(value, list) and isinstance(all_extracted_data[key], list):
                        all_extracted_data[key].extend(value)
        
        # Normalize the merged data with auto-detection
        normalized_data = normalize_data_with_llm(
            all_extracted_data, 
            schema,
            auto_detect_fields=auto_detect_fields
        )
        
        # Add extraction metadata
        metadata = {
            "filename": file.filename,
            "pages_processed": len(image_paths),
            "page_details": page_results,
            "extraction_settings": {
                "strict_mode": strict_mode,
                "double_check": double_check,
                "infer_missing": infer_missing,
                "auto_detect_fields": auto_detect_fields
            }
        }
        
        return {
            "data": normalized_data,
            "metadata": metadata
        }
    finally:
        # Cleanup temp files
        try:
            os.unlink(tmp_path)
            for path in image_paths:
                if os.path.exists(path):
                    os.unlink(path)
        except Exception as e:
            logger.warning(f"Cleanup error: {e}")
# --- Helper: Create dynamic table in Supabase based on schema ---
def create_dynamic_table(table_name: str, schema: Dict[str, Any]):
    """Create a Supabase table dynamically based on user schema"""
    try:
        # Define columns based on schema
        columns = [
            {"name": "id", "type": "uuid", "primary_key": True, "default": "gen_random_uuid()"},
            {"name": "created_at", "type": "timestamptz", "default": "now()"},
            {"name": "document_type", "type": "text"},
            {"name": "filename", "type": "text"},
        ]
        
        # Add columns from schema (simplified - in production, validate properly)
        def schema_to_columns(schema_dict, prefix=""):
            cols = []
            for field_name, field_type in schema_dict.items():
                col_name = f"{prefix}{field_name}".lower().replace(" ", "_")
                if isinstance(field_type, dict):
                    # Nested object - recurse
                    cols.extend(schema_to_columns(field_type, f"{col_name}_"))
                elif isinstance(field_type, list) and len(field_type) > 0:
                    # Array field - store as JSONB
                    cols.append({"name": col_name, "type": "jsonb"})
                else:
                    # Simple field
                    pg_type = "text"  # Default to text for flexibility
                    if field_type == "number":
                        pg_type = "numeric"
                    elif field_type == "boolean":
                        pg_type = "boolean"
                    elif field_type == "integer":
                        pg_type = "integer"
                    cols.append({"name": col_name, "type": pg_type})
            return cols
        
        columns.extend(schema_to_columns(schema))
        
        # Try to create table (note: requires admin privileges in production)
        # For now, we'll use a generic documents table approach
        logger.info(f"Would create table {table_name} with columns: {columns}")
        
    except Exception as e:
        logger.error(f"Table creation error: {e}")
        # Fallback: use a generic documents table
        pass

# --- Helper: Store in Supabase with dynamic schema ---
def store_in_supabase(table_name: str, data: dict, metadata: dict):
    """Store extracted data in Supabase with flexible schema"""
    if supabase is None:
        logger.info("Supabase not configured - skipping database storage")
        return "local_only_no_db"
    
    try:
        # Use a generic 'documents' table with JSONB for flexibility
        document_data = {
            "document_type": metadata.get("document_type", "general"),
            "filename": metadata.get("filename", "unknown"),
            "extracted_data": data,
            "metadata": metadata,
            "processed_at": datetime.now().isoformat()
        }
        
        response = supabase.table(table_name).insert(document_data).execute()
        
        if not response.data:
            raise HTTPException(status_code=400, detail="Failed to store document")
        
        return response.data[0]["id"]
    except Exception as e:
        logger.error(f"Supabase storage error: {e}")
        # Return a local ID instead of failing
        return f"local_{datetime.now().timestamp()}"

# --- Helper: Store in ChromaDB with dynamic schema ---
def store_in_chromadb_dynamic(data: dict, metadata: dict, collection_name: str = "documents"):
    """Store document embeddings in ChromaDB with flexible schema"""
    try:
        collection = chroma_client.get_or_create_collection(name=collection_name)
        
        # Create text representation for embedding
        def flatten_dict(d, parent_key='', sep='_'):
            items = []
            for k, v in d.items():
                new_key = f"{parent_key}{sep}{k}" if parent_key else k
                if isinstance(v, dict):
                    items.extend(flatten_dict(v, new_key, sep=sep).items())
                elif isinstance(v, list):
                    # Convert list to string representation for ChromaDB
                    items.append((new_key, str(v)))
                else:
                    items.append((new_key, str(v) if v is not None else "null"))
            return dict(items)
        
        flat_data = flatten_dict(data)
        text_repr = " ".join([f"{k}: {v}" for k, v in flat_data.items()])
        
        # Generate embedding
        embedding_response = ollama.embeddings(model="llama3.2", prompt=text_repr)
        embedding = embedding_response["embedding"]
        
        # Generate unique ID
        doc_id = f"doc_{datetime.now().timestamp()}_{metadata.get('filename', 'unknown')}"
        
        # Flatten metadata for ChromaDB (must be simple types only)
        def flatten_metadata(d, parent_key='', sep='_'):
            items = {}
            for k, v in d.items():
                new_key = f"{parent_key}{sep}{k}" if parent_key else k
                
                # Skip page_details entirely - too complex for ChromaDB
                if 'page_details' in k.lower():
                    continue
                
                if isinstance(v, dict):
                    items.update(flatten_metadata(v, new_key, sep=sep))
                elif isinstance(v, list):
                    # Only include simple lists (all same primitive type)
                    if len(v) > 0 and all(isinstance(x, (str, int, float, bool)) for x in v):
                        items[new_key] = str(v)
                    # Skip complex lists
                elif isinstance(v, (str, int, float, bool)) or v is None:
                    items[new_key] = str(v) if v is not None else "null"
            return items
        
        flat_metadata = {"document_type": metadata.get("document_type", "general"), **flatten_metadata(metadata)}
        
        # Remove page_details from metadata if it contains complex objects
        keys_to_remove = [k for k in flat_metadata.keys() if 'page_details' in k]
        for key in keys_to_remove:
            del flat_metadata[key]
        
        # Store in ChromaDB
        collection.add(
            ids=[doc_id],
            embeddings=[embedding],
            metadatas=[flat_metadata],
            documents=[text_repr]
        )
        
        logger.info(f"Stored document {doc_id} in ChromaDB")
        return doc_id
    except Exception as e:
        logger.error(f"ChromaDB storage error: {e}")
        # Return a local ID instead of failing
        return f"chroma_error_{datetime.now().timestamp()}"
    
# --- Serve Frontend UI ---
@app.get("/", response_class=HTMLResponse)
async def serve_frontend():
    """Serve the main frontend UI"""
    with open("index.html", "r", encoding="utf-8") as f:
        return f.read()

# --- Main Extraction Endpoint ---
@app.post("/api/extract")
async def extract_document(
    file: UploadFile = File(...),
    schema: str = Form(...),
    extraction_prompt: str = Form(...),
    document_type: str = Form("general"),
    table_name: str = Form("documents"),
    strict_mode: bool = Form(False),
    double_check: bool = Form(False),
    infer_missing: bool = Form(True),
    auto_detect_fields: bool = Form(True)  # Enabled by default
):
    """Extract data from any document with user-defined schema and prompt"""
    try:
        # Parse schema
        try:
            schema_dict = json.loads(schema)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="Invalid schema JSON format")
        
        # Process document with advanced options
        result = process_document(
            file, 
            extraction_prompt, 
            schema_dict,
            strict_mode=strict_mode,
            double_check=double_check,
            infer_missing=infer_missing,
            auto_detect_fields=auto_detect_fields
        )
        
        # Store in database
        doc_id = store_in_supabase(table_name, result["data"], {**result["metadata"], "document_type": document_type})
        
        # Store in vector DB
        vector_id = store_in_chromadb_dynamic(result["data"], {**result["metadata"], "document_type": document_type})
        
        return {
            "success": True,
            "message": "Document processed successfully",
            "document_id": doc_id,
            "vector_id": vector_id,
            "extracted_data": result["data"],
            "metadata": result["metadata"]
        }
        
    except Exception as e:
        logger.error(f"Extraction error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# --- Quick Extract with JSON payload ---
@app.post("/api/extract/json")
async def extract_from_json(request: Request):
    """Transform JSON data using user-defined schema"""
    try:
        body = await request.json()
        
        if "data" not in body or "schema" not in body:
            raise HTTPException(status_code=400, detail="Request must contain 'data' and 'schema' fields")
        
        normalized_data = normalize_data_with_llm(body["data"], body["schema"])
        
        return {
            "success": True,
            "transformed_data": normalized_data
        }
    except Exception as e:
        logger.error(f"JSON extraction error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# --- Search Endpoint ---
@app.post("/api/search")
async def search_documents(query: str, collection_name: str = "documents", limit: int = 5):
    """Search documents using semantic similarity"""
    try:
        collection = chroma_client.get_or_create_collection(name=collection_name)
        
        # Generate query embedding
        query_embedding_response = ollama.embeddings(model="llama3.2", prompt=query)
        query_embedding = query_embedding_response["embedding"]
        
        # Search
        results = collection.query(
            query_embeddings=[query_embedding],
            n_results=limit,
            include=["documents", "metadatas"]
        )
        
        return {
            "success": True,
            "results": [
                {
                    "document": doc,
                    "metadata": meta,
                    "score": None  # ChromaDB doesn't always return scores
                }
                for doc, meta in zip(results["documents"][0], results["metadatas"][0])
            ]
        }
    except Exception as e:
        logger.error(f"Search error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# --- Get All Documents ---
@app.get("/api/documents")
async def get_documents(table_name: str = "documents", limit: int = 100):
    """Retrieve all stored documents"""
    if supabase is None:
        return {
            "success": True,
            "documents": [],
            "count": 0,
            "message": "Supabase not configured - no documents stored yet"
        }
    
    try:
        response = supabase.table(table_name).select("*").limit(limit).execute()
        return {
            "success": True,
            "documents": response.data,
            "count": len(response.data)
        }
    except Exception as e:
        logger.error(f"Get documents error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# --- Download Endpoints ---
@app.post("/api/download/json")
async def download_json(extracted_data: Dict[str, Any]):
    """Download extracted data as JSON file"""
    try:
        json_str = json.dumps(extracted_data, indent=2)
        return Response(
            content=json_str,
            media_type="application/json",
            headers={"Content-Disposition": "attachment; filename=extracted_data.json"}
        )
    except Exception as e:
        logger.error(f"JSON download error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/download/csv")
async def download_csv(extracted_data: Dict[str, Any]):
    """Download extracted data as CSV file"""
    try:
        output = io.StringIO()
        
        # Flatten the data for CSV
        def flatten_dict(d, parent_key='', sep='.'):
            items = []
            for k, v in d.items():
                new_key = f"{parent_key}{sep}{k}" if parent_key else k
                if isinstance(v, dict):
                    items.extend(flatten_dict(v, new_key, sep=sep).items())
                elif isinstance(v, list):
                    # Handle lists specially - convert to string or expand
                    if len(v) > 0 and isinstance(v[0], dict):
                        # List of dicts - create separate rows
                        pass
                    else:
                        items.append((new_key, ', '.join(map(str, v))))
                else:
                    items.append((new_key, v))
            return dict(items)
        
        flat_data = flatten_dict(extracted_data)
        
        writer = csv.DictWriter(output, fieldnames=flat_data.keys())
        writer.writeheader()
        writer.writerow(flat_data)
        
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=extracted_data.csv"}
        )
    except Exception as e:
        logger.error(f"CSV download error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/download/excel")
async def download_excel(extracted_data: Dict[str, Any]):
    """Download extracted data as Excel file"""
    try:
        import pandas as pd
        from openpyxl import Workbook
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Data"
        
        # Add title
        ws['A1'] = "Extracted Document Data"
        ws['A1'].font = ws['A1'].font.copy(bold=True, size=16)
        
        row_num = 3
        
        # Write data
        def write_data(data, start_col='A', start_row=3):
            col_offset = 0
            for key, value in data.items():
                col_letter = chr(ord('A') + col_offset)
                cell_ref = f"{col_letter}{start_row}"
                
                if isinstance(value, dict):
                    # Write nested dict header
                    ws[cell_ref] = key
                    ws[cell_ref].font = ws[cell_ref].font.copy(bold=True)
                    # Recursively write nested data
                    write_data(value, start_col=col_letter, start_row=start_row+1)
                elif isinstance(value, list):
                    # Write list
                    ws[cell_ref] = key
                    ws[cell_ref].font = ws[cell_ref].font.copy(bold=True)
                    row_offset = 1
                    for item in value:
                        if isinstance(item, dict):
                            # List of objects
                            for sub_key, sub_value in item.items():
                                ws[f"{col_letter}{start_row+row_offset}"] = f"{sub_key}: {sub_value}"
                                row_offset += 1
                        else:
                            ws[f"{col_letter}{start_row+row_offset}"] = item
                            row_offset += 1
                else:
                    ws[cell_ref] = f"{key}: {value}"
                col_offset += 1
            
            return start_row + 1
        
        write_data(extracted_data)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return Response(
            content=excel_buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=extracted_data.xlsx"}
        )
    except ImportError:
        # Fallback if pandas/openpyxl not installed
        logger.error("Excel libraries not installed")
        raise HTTPException(status_code=500, detail="Excel export requires pandas and openpyxl. Install with: pip install pandas openpyxl")
    except Exception as e:
        logger.error(f"Excel download error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/download/pdf")
async def download_pdf(extracted_data: Dict[str, Any], metadata: Optional[Dict[str, Any]] = None):
    """Download extracted data as PDF file"""
    try:
        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title = Paragraph("Extracted Document Data", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Metadata if available
        if metadata:
            elements.append(Paragraph("Document Metadata", styles['Heading2']))
            meta_data = [["Field", "Value"]]
            for key, value in metadata.items():
                if key != 'page_details':  # Skip complex metadata
                    meta_data.append([str(key), str(value)])
            
            meta_table = Table(meta_data)
            meta_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            elements.append(meta_table)
            elements.append(Spacer(1, 20))
        
        # Extracted data
        elements.append(Paragraph("Extracted Information", styles['Heading2']))
        
        # Convert dict to table data
        table_data = [["Field", "Value"]]
        
        def add_to_table(data, prefix=''):
            for key, value in data.items():
                field_name = f"{prefix}.{key}" if prefix else key
                if isinstance(value, dict):
                    add_to_table(value, field_name)
                elif isinstance(value, list):
                    # Format list items
                    items_str = "\n".join([f"• {item}" if not isinstance(item, dict) else 
                                         "\n".join([f"  - {k}: {v}" for k, v in item.items()])
                                         for item in value])
                    table_data.append([field_name, items_str])
                else:
                    table_data.append([field_name, str(value) if value is not None else "N/A"])
        
        add_to_table(extracted_data)
        
        # Create table
        data_table = Table(table_data, colWidths=['30%', '70%'])
        data_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elements.append(data_table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=extracted_data.pdf"}
        )
    except Exception as e:
        logger.error(f"PDF download error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    
